
import openpyxl
import csv
import os

path = 'home/isocyanate/excelToCSV_testing'

for i in os.listdir(path):
    if i.endswith('xlsx'):
        filename = os.path.basename(i)
        filename = filename.replace('.xlsx', '')
        wb = openpyxl.load_workbook(i)

        for j in wb.worksheets:
            csvName = filename + '_' + str(j.title) + '.csv'    
            csvFile = open('home/isocyanate/excelToCSV_testing'+csvName, 'w', newline='')
            csvWriter = csv.writer(csvFile)

        for k in range(1,i.max_row+1):
            listy = []
            for l in range(1,i.max_column+1):
                listy.append(i.cell(row=k,column=l).value)
            csvWriter.writerow(listy)
        csvFile.close()
