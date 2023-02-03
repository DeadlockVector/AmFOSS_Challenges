import openpyxl, sys

#testing
filename = '/home/isocyanate/Desktop/AmFOSS_Challenges/Chapter13/BlankRow.xlsx'
row_number = 3
number_blanks = 2

wb = openpyxl.load_workbook(filename)
sheet = wb.active

wb1 = openpyxl.Workbook()
new_sheet = wb1.active

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        if i <= row_number:
            new_sheet.cell(row=i, column=j).value= sheet.cell(row=i, column=j).value
        elif i > row_number:
            new_sheet.cell(row=i+number_blanks, columns=j).value = sheet.cell(row=i, column=j).value

wb1.save('/home/isocyanate/Desktop/AmFOSS_Challenges/Chapter13/new_excel.xlsx')
wb.save('/home/isocyanate/Desktop/AmFOSS_Challenges/Chapter13/insertBlank.xlsx')

'''
wb = openpyxl.Workbook()
sheet = wb.active
wb.save('/home/isocyanate/Desktop/AmFOSS_Challenges/Chapter13/BlankRow.xlsx')
'''