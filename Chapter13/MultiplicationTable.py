import openpyxl
from openpyxl.styles import Font

#n = int(input('Enter the multiplication number - '))
#testing with number 5
multiplication_number = 5

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, multiplication_number+2):
    for j in range(1, multiplication_number+2):
        if i==j==1:
            sheet.cell(row=i, column=j).value = ''
        elif i==1 or j==1:
            sheet.cell(row=i, column=j).value = (i)*(j)
        else:
            sheet.cell(row=i, column=j).value = (i-1)*(j-1)

for i in range(1, multiplication_number+2):
    for j in range(1, multiplication_number+2):
        if i==1:
            sheet.cell(row=1, column=j).font = Font(bold=True)
        elif j==1:
            sheet.cell(row=i, column=1).font = Font(bold = True)


wb.save('/home/isocyanate/Desktop/AmFOSS_Challenges/Chapter13/MultiplicationTable.xlsx')