import openpyxl

file1 = '/home/isocyanate/Desktop/AmFOSS_Challenges/Chapter13/SpreadsheetToText/Testing.xlsx'
wb = openpyxl.load_workbook(file1)
sheet = wb.active

number_files = sheet.max_column
number_characters = sheet.max_row
list_files = []

for i in range(1, number_files+1):
    list_files.append("/home/isocyanate/Desktop/AmFOSS_Challenges/Chapter13/SpreadsheetToText/testing"+str(i)+'.txt')

for i in range(1, number_files+1):
    for j in range(1, number_characters+1):
        #opening each file in append mode
        files = open(list_files[i-1], 'a')

        files.write(sheet.cell(row=j, column=i).value)
        
    print()
