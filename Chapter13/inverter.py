
import openpyxl

filename = "/home/isocyanate/Desktop/AmFOSS_Challenges/Chapter13/InvertingTester.xlsx"

wb = openpyxl.load_workbook(filename)
sheet = wb.active

wb1 = openpyxl.Workbook()
sheet1 = wb1.active

for i in range(1, sheet.max_row):
    for j in range(1, sheet.max_column):
        sheet1.cell(row = j, column = i).value = sheet.cell(row = i, column = j).value

wb.save("/home/isocyanate/Desktop/AmFOSS_Challenges/Chapter13/Inverted.xlsx")
wb1.save("/home/isocyanate/Desktop/AmFOSS_Challenges/Chapter13/Inverted.xlsx")

'''
wb = openpyxl.Workbook()
sheet = wb.active
wb.save('/home/isocyanate/Desktop/AmFOSS_Challenges/Chapter13/InvertingTester.xlsx')
'''